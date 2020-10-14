# import xlrd
# from xlutils.copy import copy
import openpyxl


class OperaExcel:

    def __init__(self, file_path=None, sheet_index=None):
        if file_path is None:
            self.file_path = '../config/case/case1.xlsx'
        else:
            self.file_path = file_path
        self.wb = self.get_excel_workbook()
        if sheet_index is None:
            self.sheet_index = 0
        else:
            self.sheet_index = sheet_index
        self.ws = self.get_excel_worksheet()
        self.max_row = self.ws.max_row
        self.max_col = self.ws.max_column

    def get_excel_workbook(self):
        """
        获取excel所有内容
        :return:
        """
        # excel_data = xlrd.open_workbook(self.file_path) # xlrd模块 打开workbook
        wb = openpyxl.load_workbook(self.file_path)
        return wb

    def get_excel_worksheet(self):
        """
        获取sheets内容
        :param i:
        :return:
        """
        try:
            # sheet = self.excel.sheets()[self.sheet_num]   # xlrd模块 读取sheet
            ws = self.wb['Sheet1']  # 通过sheet名称打开( 注意大小写区分）
            # sheet_names = self.wb.get_sheet_names()     # 不知道sheet名称时，可通过index下标获取
            # ws = self.wb.get_sheet_by_name(sheet_names[0])
        except IndexError:
            ws = None

        return ws

    # def get_lines(self):
    #     """
    #     获取sheet总行数
    #     :return:
    #     """
    #     lines = self.data.nrows   # xlrd模块 获取sheet总行数
    #     return lines

    # def get_cols(self):
    #     cols = self.data.ncols   # xlrd模块 获取sheet总列数
    #     return cols

    def get_value(self, row, col):
        """
        获取单元格的值
        :return:
        """
        if row > self.max_row or col > self.max_col:
            value = None
        else:
            # data = self.ws.cell(row, col).value  # xlrd模块 获取单元格的值
            value = self.ws.cell(row=row, column=col).value  # 可通过行和列值直接获取，注意openpyxl行和列是从1开始的
            # value = self.ws['B1'].value  # openpyxl 也可通过单元格名称直接获取值
        return value

    def write_excel(self, row, col, value):
        self.ws.cell(row=row, column=col).value = value
        self.wb.save(self.file_path)
        pass


if __name__ == '__main__':
    exc = OperaExcel()
    exc.write_excel(2, 4, 'goodbye2')
    print(exc.get_value(2, 2))
# print(exc.get_value(1, 4))
